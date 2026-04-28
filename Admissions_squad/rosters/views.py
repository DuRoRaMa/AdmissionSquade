from io import BytesIO
from urllib.parse import quote
from django.http import HttpResponse
from django.db.models import Count
from collections import defaultdict
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.utils import timezone
from django.shortcuts import get_object_or_404

from rest_framework import generics, status
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from squads.models import SquadMembership
from .models import (
    AvailabilityForm,
    WorkBlock,
    SquadMembership,
    Schedule,
    ScheduleEntry,
    ScheduleChangeRequest,
    QrToken,
    AvailabilitySlot
)
from .serializers import (
    AvailabilityResponseMemberSerializer,
    AvailabilityFormSerializer,
    AvailabilitySubmitSerializer,
    ScheduleSerializer,
    ScheduleEntrySerializer,
    ScheduleChangeRequestSerializer,
    WorkBlockSerializer,
)
from .permissions import (
    can_manage_availability,
    can_respond_own_availability,
    can_manage_roster,
    can_publish_roster,
    can_view_roster_all,
    can_view_own_roster,
    get_squad_ids_with_permission,
    get_user_active_memberships,
)
from .services.availability import submit_availability
from .services.generator import generate_schedule
from .services.changes import approve_change_request, reject_change_request
from .services.attendance import create_qr_token, scan_qr


class AvailabilityFormListCreateView(generics.ListCreateAPIView):
    serializer_class = AvailabilityFormSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = (
            AvailabilityForm.objects
            .all()
            .prefetch_related("days__shifts")
            .select_related("squad", "created_by")
        )

        if self.request.user.is_staff:
            return queryset

        squad_ids = get_squad_ids_with_permission(
            self.request.user,
            ["availability.manage"],
        )
        return queryset.filter(squad_id__in=squad_ids)

    def perform_create(self, serializer):
        squad = serializer.validated_data["squad"]

        if not can_manage_availability(self.request.user, squad):
            raise PermissionDenied("Недостаточно прав для создания формы доступности.")

        serializer.save(created_by=self.request.user)


class AvailabilityFormOpenView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        form = get_object_or_404(AvailabilityForm, pk=pk)

        if not can_manage_availability(request.user, form.squad):
            raise PermissionDenied("Недостаточно прав для открытия формы доступности.")

        form.status = "open"
        form.save(update_fields=["status"])
        return Response({"message": "Форма открыта"})


class AvailabilityFormCloseView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        form = get_object_or_404(AvailabilityForm, pk=pk)

        if not can_manage_availability(request.user, form.squad):
            raise PermissionDenied("Недостаточно прав для закрытия формы доступности.")

        form.status = "closed"
        form.save(update_fields=["status"])
        return Response({"message": "Форма закрыта"})

class WorkBlockListCreateView(generics.ListCreateAPIView):
    serializer_class = WorkBlockSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = WorkBlock.objects.filter(is_active=True).select_related("squad")

        squad_id = self.request.query_params.get("squad")

        if squad_id:
            queryset = queryset.filter(squad_id=squad_id)

        if self.request.user.is_staff:
            return queryset

        squad_ids = get_squad_ids_with_permission(
            self.request.user,
            ["roster.view_all", "roster.manage", "roster.publish"],
        )

        return queryset.filter(squad_id__in=squad_ids)

    def perform_create(self, serializer):
        squad = serializer.validated_data["squad"]

        if not can_manage_roster(self.request.user, squad):
            raise PermissionDenied("Недостаточно прав для создания блока работ.")

        serializer.save()

class ActiveAvailabilityFormView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        squad_id = request.query_params.get("squad")

        memberships = get_user_active_memberships(request.user)
        if squad_id:
            memberships = memberships.filter(squad_id=squad_id)

        allowed_memberships = [
            membership
            for membership in memberships
            if can_respond_own_availability(request.user, membership.squad)
        ]

        if not allowed_memberships:
            return Response(
                {"detail": "У вас нет активного доступа к форме доступности."},
                status=404,
            )

        squad = allowed_memberships[0].squad
        form = (
            AvailabilityForm.objects
            .filter(squad=squad, status="open")
            .prefetch_related("days__shifts")
            .order_by("-created_at")
            .first()
        )

        if not form:
            return Response({"detail": "Нет открытой формы доступности."}, status=404)

        serializer = AvailabilityFormSerializer(form)
        return Response(serializer.data)


class SubmitAvailabilityView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        form = get_object_or_404(AvailabilityForm, pk=pk)

        membership = SquadMembership.objects.filter(
            user=request.user,
            squad=form.squad,
            is_active=True,
        ).first()

        if not membership:
            return Response({"detail": "Вы не состоите в этом отряде."}, status=403)

        if not can_respond_own_availability(request.user, form.squad):
            raise PermissionDenied("Недостаточно прав для отправки доступности.")

        if form.status != "open":
            return Response({"detail": "Форма не открыта."}, status=400)

        if form.response_deadline and timezone.now() > form.response_deadline:
            return Response({"detail": "Срок отправки формы истёк."}, status=400)

        serializer = AvailabilitySubmitSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            submit_availability(form, membership, serializer.validated_data["slots"])
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)

        return Response({"message": "Ответ принят"})


class ScheduleListCreateView(generics.ListCreateAPIView):
    serializer_class = ScheduleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = (
            Schedule.objects
            .select_related("squad", "availability_form", "created_by")
            .prefetch_related("needs", "needs__work_block")
            .annotate(entries_count_value=Count("entries"))
        )

        if self.request.user.is_staff:
            return queryset

        squad_ids = get_squad_ids_with_permission(
            self.request.user,
            ["roster.view_all", "roster.manage", "roster.publish"],
        )

        return queryset.filter(squad_id__in=squad_ids)

    def perform_create(self, serializer):
        squad = serializer.validated_data["squad"]

        if not can_manage_roster(self.request.user, squad):
            raise PermissionDenied("Недостаточно прав для создания графика.")

        serializer.save(created_by=self.request.user)


class GenerateScheduleView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        schedule = get_object_or_404(
            Schedule.objects.select_related("squad", "availability_form"),
            pk=pk,
        )

        if not can_manage_roster(request.user, schedule.squad):
            raise PermissionDenied("Недостаточно прав для генерации графика.")

        if schedule.status != "draft":
            return Response(
                {"detail": "Генерировать можно только черновик."},
                status=400,
            )

        try:
            generate_schedule(schedule)
        except ValueError as e:
            return Response(
                {"detail": str(e)},
                status=400,
            )
        entries_count = schedule.entries.count()
        return Response({
            "message": "Черновик графика сформирован.",
            "schedule_id": schedule.id,
            "has_entries": entries_count > 0,
            "entries_count": entries_count,
        })

class ScheduleEntriesView(generics.ListAPIView):
    serializer_class = ScheduleEntrySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        schedule = get_object_or_404(
            Schedule.objects.select_related("squad"),
            pk=self.kwargs["pk"],
        )

        if not (
            can_manage_roster(self.request.user, schedule.squad)
            or can_view_roster_all(self.request.user, schedule.squad)
        ):
            raise PermissionDenied("Недостаточно прав для просмотра графика.")

        return (
            ScheduleEntry.objects
            .filter(schedule=schedule)
            .select_related(
                "schedule",
                "need",
                "membership",
                "membership__user",
                "membership__role",
                "work_block",
            )
            .order_by(
                "date",
                "starts_at",
                "work_block__name",
                "membership__user__last_name",
                "membership__user__first_name",
            )
        )

class PublishScheduleView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        schedule = get_object_or_404(Schedule, pk=pk)

        if not can_publish_roster(request.user, schedule.squad):
            raise PermissionDenied("Недостаточно прав для публикации графика.")

        if schedule.status != "draft":
            return Response({"detail": "Опубликовать можно только черновик."}, status=400)

        schedule.status = "published"
        schedule.published_at = timezone.now()
        schedule.save(update_fields=["status", "published_at"])

        return Response({"message": "График опубликован"})


class MyScheduleView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.is_staff:
            queryset = (
                ScheduleEntry.objects
                .filter(schedule__status="published")
                .select_related("work_block", "schedule", "membership__user", "membership__squad")
            )
            serializer = ScheduleEntrySerializer(queryset, many=True)
            return Response(serializer.data)

        allowed_membership_ids = [
            membership.id
            for membership in get_user_active_memberships(request.user)
            if can_view_own_roster(request.user, membership.squad)
        ]

        if not allowed_membership_ids:
            return Response([], status=200)

        entries = (
            ScheduleEntry.objects
            .filter(
                membership_id__in=allowed_membership_ids,
                schedule__status="published",
            )
            .select_related("work_block", "schedule", "membership__user", "membership__squad")
        )

        serializer = ScheduleEntrySerializer(entries, many=True)
        return Response(serializer.data)


class ChangeRequestCreateView(generics.CreateAPIView):
    serializer_class = ScheduleChangeRequestSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)


class MyChangeRequestsView(generics.ListAPIView):
    serializer_class = ScheduleChangeRequestSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return (
            ScheduleChangeRequest.objects
            .filter(requested_by=self.request.user)
            .select_related(
                "entry",
                "entry__schedule",
                "entry__membership",
                "entry__membership__user",
                "target_membership",
                "reviewed_by",
            )
            .order_by("-created_at")
        )


class ChangeRequestListView(generics.ListAPIView):
    serializer_class = ScheduleChangeRequestSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = (
            ScheduleChangeRequest.objects
            .all()
            .select_related(
                "entry",
                "entry__schedule",
                "entry__schedule__squad",
                "entry__membership",
                "requested_by",
                "target_membership",
                "reviewed_by",
            )
            .order_by("-created_at")
        )

        if self.request.user.is_staff:
            return queryset

        squad_ids = get_squad_ids_with_permission(self.request.user, "roster.manage")
        return queryset.filter(entry__schedule__squad_id__in=squad_ids)


class ApproveChangeRequestView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        change_request = get_object_or_404(
            ScheduleChangeRequest.objects.select_related("entry__schedule__squad"),
            pk=pk,
        )

        if not can_manage_roster(request.user, change_request.entry.schedule.squad):
            raise PermissionDenied("Недостаточно прав для одобрения заявки.")

        try:
            approve_change_request(change_request, request.user)
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)

        return Response({"message": "Заявка одобрена"})


class RejectChangeRequestView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        change_request = get_object_or_404(
            ScheduleChangeRequest.objects.select_related("entry__schedule__squad"),
            pk=pk,
        )

        if not can_manage_roster(request.user, change_request.entry.schedule.squad):
            raise PermissionDenied("Недостаточно прав для отклонения заявки.")

        reject_change_request(
            change_request,
            request.user,
            request.data.get("review_comment", ""),
        )
        return Response({"message": "Заявка отклонена"})


class CreateQrTokenView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, entry_id):
        entry = get_object_or_404(
            ScheduleEntry.objects.select_related("membership__user"),
            pk=entry_id,
        )

        if entry.membership.user_id != request.user.id and not request.user.is_staff:
            return Response({"detail": "Это не ваша смена."}, status=403)

        token = create_qr_token(entry)
        return Response(
            {
                "message": "QR-токен создан",
                "token": token.token,
                "expires_at": token.expires_at,
            }
        )


class ScanQrView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        token_value = request.data.get("token")
        if not token_value:
            return Response({"detail": "Токен обязателен."}, status=400)

        token = get_object_or_404(
            QrToken.objects.select_related("entry__schedule__squad"),
            token=token_value,
        )

        if not can_manage_roster(request.user, token.entry.schedule.squad):
            raise PermissionDenied("Недостаточно прав для отметки прихода или ухода.")

        try:
            result = scan_qr(token_value, request.user)
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)

        return Response(result)
class AvailabilityFormResponsesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        form = get_object_or_404(
            AvailabilityForm.objects.select_related("squad"),
            pk=pk,
        )
        if not can_manage_availability(request.user, form.squad):
            raise PermissionDenied("Недостаточно прав для просмотра ответов на форму.")

        memberships = list(
            SquadMembership.objects.filter(
                squad=form.squad,
                is_active=True,
            )
            .select_related("user", "role")
            .order_by("user__last_name", "user__first_name", "user__middle_name")
        )

        slots = (
            AvailabilitySlot.objects.filter(
                shift__day__form=form,
                membership__in=memberships,
            )
            .select_related("membership__user", "membership__role", "shift__day")
            .order_by("membership_id", "shift__day__date", "shift__starts_at")
        )

        slots_by_membership = {}
        for slot in slots:
            slots_by_membership.setdefault(slot.membership_id, []).append(
                {
                    "shift_id": slot.shift_id,
                    "date": slot.shift.day.date,
                    "shift_title": slot.shift.title,
                    "starts_at": slot.shift.starts_at,
                    "ends_at": slot.shift.ends_at,
                    "is_available": slot.is_available,
                    "comment": slot.comment or "",
                    "submitted_at": slot.submitted_at,
                }
            )

        members_payload = []
        for membership in memberships:
            user = membership.user
            member_slots = slots_by_membership.get(membership.id, [])

            submitted_at = None
            if member_slots:
                submitted_values = [item["submitted_at"] for item in member_slots if item["submitted_at"]]
                submitted_at = max(submitted_values) if submitted_values else None

            members_payload.append(
                {
                    "membership_id": membership.id,
                    "user_id": user.id,
                    "full_name": f"{user.last_name} {user.first_name} {user.middle_name}".strip() or user.email,
                    "role_name": membership.role.name if membership.role else "",
                    "has_response": bool(member_slots),
                    "available_count": sum(1 for item in member_slots if item["is_available"]),
                    "unavailable_count": sum(1 for item in member_slots if not item["is_available"]),
                    "submitted_at": submitted_at,
                    "slots": member_slots,
                }
            )

        serializer = AvailabilityResponseMemberSerializer(members_payload, many=True)

        return Response(
            {
                "form": {
                    "id": form.id,
                    "title": form.title,
                    "status": form.status,
                    "squad": form.squad_id,
                    "period_start": form.period_start,
                    "period_end": form.period_end,
                    "response_deadline": form.response_deadline,
                },
                "members": serializer.data,
            },
            status=status.HTTP_200_OK,
        )

class AvailabilityFormResponsesExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        form = get_object_or_404(
            AvailabilityForm.objects
            .select_related("squad")
            .prefetch_related("days__shifts"),
            pk=pk,
        )

        if not can_manage_availability(request.user, form.squad):
            raise PermissionDenied("Недостаточно прав для экспорта ответов на форму.")

        memberships = list(
            SquadMembership.objects
            .filter(squad=form.squad, is_active=True)
            .select_related("user", "role")
            .order_by("user__last_name", "user__first_name", "user__middle_name")
        )

        shifts = []
        for day in form.days.all().order_by("date"):
            for shift in day.shifts.filter(is_active=True).order_by("starts_at", "id"):
                shifts.append(shift)

        slots = (
            AvailabilitySlot.objects
            .filter(
                shift__day__form=form,
                membership__in=memberships,
            )
            .select_related("membership", "shift", "shift__day")
        )

        slots_by_key = {
            (slot.membership_id, slot.shift_id): slot
            for slot in slots
        }

        responded_membership_ids = {
            slot.membership_id
            for slot in slots
        }

        wb = Workbook()
        primary_ws = wb.active
        primary_ws.title = "Основные смены"

        extra_ws = wb.create_sheet("Доп. смены")

        primary_shifts = [
            shift for shift in shifts
            if shift.shift_kind == "primary"
        ]

        extra_shifts = [
            shift for shift in shifts
            if shift.shift_kind == "extra"
        ]

        self._fill_sheet(
            ws=primary_ws,
            form=form,
            memberships=memberships,
            shifts=primary_shifts,
            slots_by_key=slots_by_key,
            responded_membership_ids=responded_membership_ids,
            sheet_title="Основные смены",
        )

        self._fill_sheet(
            ws=extra_ws,
            form=form,
            memberships=memberships,
            shifts=extra_shifts,
            slots_by_key=slots_by_key,
            responded_membership_ids=responded_membership_ids,
            sheet_title="Дополнительные смены",
        )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_title = "".join(
            char if char.isalnum() or char in (" ", "_", "-") else "_"
            for char in form.title
        ).strip()

        filename = f"{safe_title or 'availability_form'}_{form.id}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{quote(filename)}"
        )

        return response

    def _fill_sheet(
        self,
        ws,
        form,
        memberships,
        shifts,
        slots_by_key,
        responded_membership_ids,
        sheet_title,
    ):
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        header_fill = PatternFill("solid", fgColor="F2F2F2")
        counter_fill = PatternFill("solid", fgColor="EAF2F8")

        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        bold = Font(bold=True)
        normal = Font(bold=False)

        last_col = max(2, len(shifts) + 1)

        ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=1)
        fio_cell = ws.cell(row=1, column=1, value="ФИО")
        fio_cell.font = bold
        fio_cell.alignment = center
        fio_cell.fill = header_fill

        self._style_range(ws, 1, 1, 4, 1, border, header_fill, center, bold)

        if shifts:
            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)

            date_title_cell = ws.cell(row=1, column=2, value="Дата")
            date_title_cell.font = bold
            date_title_cell.alignment = center
            date_title_cell.fill = header_fill

            self._style_range(ws, 1, 2, 1, last_col, border, header_fill, center, bold)

            col = 2
            index = 0

            while index < len(shifts):
                current_date = shifts[index].day.date
                same_date_shifts = []

                while (
                    index + len(same_date_shifts) < len(shifts)
                    and shifts[index + len(same_date_shifts)].day.date == current_date
                ):
                    same_date_shifts.append(shifts[index + len(same_date_shifts)])

                start_col = col
                end_col = col + len(same_date_shifts) - 1

                if start_col != end_col:
                    ws.merge_cells(
                        start_row=2,
                        start_column=start_col,
                        end_row=2,
                        end_column=end_col,
                    )
                    ws.merge_cells(
                        start_row=3,
                        start_column=start_col,
                        end_row=3,
                        end_column=end_col,
                    )

                date_cell = ws.cell(row=2, column=start_col, value=current_date)
                date_cell.number_format = "dd.mm.yyyy"
                date_cell.font = bold
                date_cell.alignment = center
                date_cell.fill = header_fill

                available_count = self._count_available_members_for_day(
                    memberships=memberships,
                    shifts=same_date_shifts,
                    slots_by_key=slots_by_key,
                )

                counter_cell = ws.cell(
                    row=3,
                    column=start_col,
                    value=f"Хочет выйти: {available_count}",
                )
                counter_cell.font = bold
                counter_cell.alignment = center
                counter_cell.fill = counter_fill

                self._style_range(
                    ws,
                    2,
                    start_col,
                    2,
                    end_col,
                    border,
                    header_fill,
                    center,
                    bold,
                )

                self._style_range(
                    ws,
                    3,
                    start_col,
                    3,
                    end_col,
                    border,
                    counter_fill,
                    center,
                    bold,
                )

                for offset, shift in enumerate(same_date_shifts):
                    shift_title = shift.title or shift.get_shift_kind_display()
                    starts_at = shift.starts_at.strftime("%H:%M") if shift.starts_at else ""
                    ends_at = shift.ends_at.strftime("%H:%M") if shift.ends_at else ""

                    title = shift_title

                    if starts_at and ends_at:
                        title = f"{shift_title}\n{starts_at}–{ends_at}"

                    shift_cell = ws.cell(row=4, column=col + offset, value=title)
                    shift_cell.font = bold
                    shift_cell.alignment = center
                    shift_cell.fill = header_fill
                    shift_cell.border = border

                col += len(same_date_shifts)
                index += len(same_date_shifts)
        else:
            ws.merge_cells(start_row=1, start_column=2, end_row=4, end_column=2)
            empty_cell = ws.cell(row=1, column=2, value=f"{sheet_title}: нет активных смен")
            empty_cell.font = bold
            empty_cell.alignment = center
            empty_cell.fill = header_fill
            self._style_range(ws, 1, 2, 4, 2, border, header_fill, center, bold)

        for row_index, membership in enumerate(memberships, start=5):
            user = membership.user
            middle_name = getattr(user, "middle_name", "") or ""

            full_name = (
                f"{user.last_name} {user.first_name} {middle_name}".strip()
                or user.email
            )

            fio_cell = ws.cell(row=row_index, column=1, value=full_name)
            fio_cell.alignment = left
            fio_cell.font = normal
            fio_cell.border = border

            for col_index, shift in enumerate(shifts, start=2):
                cell = ws.cell(row=row_index, column=col_index)

                slot = slots_by_key.get((membership.id, shift.id))

                if membership.id not in responded_membership_ids:
                    value = "-"
                elif slot is None:
                    value = "-"
                elif slot.is_available:
                    value = 1
                else:
                    value = 0

                cell.value = value
                cell.alignment = center
                cell.border = border

        max_row = max(5, len(memberships) + 4)

        for row in range(1, max_row + 1):
            for col in range(1, last_col + 1):
                ws.cell(row=row, column=col).border = border

        ws.column_dimensions["A"].width = 34

        for col in range(2, last_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 24
        ws.row_dimensions[3].height = 24
        ws.row_dimensions[4].height = 42

        for row in range(5, max_row + 1):
            ws.row_dimensions[row].height = 22

        ws.freeze_panes = "B5"

    def _count_available_members_for_day(self, memberships, shifts, slots_by_key):
        available_members = set()

        shift_ids = [
            shift.id
            for shift in shifts
        ]

        for membership in memberships:
            for shift_id in shift_ids:
                slot = slots_by_key.get((membership.id, shift_id))

                if slot and slot.is_available:
                    available_members.add(membership.id)
                    break

        return len(available_members)

    def _style_range(
        self,
        ws,
        start_row,
        start_column,
        end_row,
        end_column,
        border,
        fill,
        alignment,
        font,
    ):
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.fill = fill
                cell.alignment = alignment
                cell.font = font

class ScheduleExportXlsxView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        schedule = get_object_or_404(
            Schedule.objects.select_related("squad"),
            pk=pk,
        )

        if not (
            can_manage_roster(request.user, schedule.squad)
            or can_view_roster_all(request.user, schedule.squad)
        ):
            raise PermissionDenied("Недостаточно прав для скачивания графика.")

        entries = list(
            schedule.entries
            .exclude(status="cancelled")
            .select_related(
                "membership",
                "membership__user",
                "work_block",
            )
            .order_by(
                "date",
                "starts_at",
                "ends_at",
                "work_block__name",
                "membership__user__last_name",
                "membership__user__first_name",
            )
        )

        if not entries:
            return Response(
                {"detail": "График ещё не сгенерирован."},
                status=400,
            )

        memberships = list(
            SquadMembership.objects
            .filter(squad=schedule.squad, is_active=True)
            .select_related("user")
            .order_by(
                "user__last_name",
                "user__first_name",
                "user__middle_name",
                "id",
            )
        )

        dates = self.get_schedule_dates(schedule, entries)

        entries_by_member_and_date = defaultdict(list)
        people_count_by_date = defaultdict(set)

        for entry in entries:
            entries_by_member_and_date[
                (entry.membership_id, entry.date)
            ].append(entry)

            people_count_by_date[entry.date].add(entry.membership_id)

        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Сводка"

        self.build_summary_header(
            sheet=summary_sheet,
            dates=dates,
            people_count_by_date=people_count_by_date,
        )

        self.fill_summary_body(
            sheet=summary_sheet,
            memberships=memberships,
            dates=dates,
            entries_by_member_and_date=entries_by_member_and_date,
        )

        self.apply_summary_styles(
            sheet=summary_sheet,
            dates=dates,
        )

        for day in dates:
            day_sheet = workbook.create_sheet(
                title=self.get_safe_sheet_title(day.strftime("%d.%m.%Y")),
            )

            self.build_day_sheet(
                sheet=day_sheet,
                day=day,
                memberships=memberships,
                entries_by_member_and_date=entries_by_member_and_date,
                people_count=len(people_count_by_date.get(day, set())),
            )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"График_{schedule.title}.xlsx"
        encoded_filename = quote(filename)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{encoded_filename}"
        )

        return response

    def get_schedule_dates(self, schedule, entries):
        if schedule.period_start and schedule.period_end:
            dates = []
            current = schedule.period_start

            while current <= schedule.period_end:
                dates.append(current)
                current += timedelta(days=1)

            return dates

        return sorted({entry.date for entry in entries})

    def build_summary_header(self, sheet, dates, people_count_by_date):
        last_column = 1 + len(dates) * 2

        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=4,
            end_column=1,
        )
        sheet.cell(row=1, column=1, value="ФИО")

        sheet.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=last_column,
        )
        sheet.cell(row=1, column=2, value="День")

        column = 2

        for day in dates:
            sheet.merge_cells(
                start_row=2,
                start_column=column,
                end_row=2,
                end_column=column + 1,
            )

            date_cell = sheet.cell(row=2, column=column, value=day)
            date_cell.number_format = "DD.MM.YYYY"

            sheet.merge_cells(
                start_row=3,
                start_column=column,
                end_row=3,
                end_column=column + 1,
            )

            sheet.cell(
                row=3,
                column=column,
                value=len(people_count_by_date.get(day, set())),
            )

            sheet.cell(row=4, column=column, value="Выхожу")
            sheet.cell(row=4, column=column + 1, value="Блок")

            column += 2

    def fill_summary_body(
        self,
        sheet,
        memberships,
        dates,
        entries_by_member_and_date,
    ):
        row = 5

        for membership in memberships:
            sheet.cell(
                row=row,
                column=1,
                value=self.get_member_name(membership),
            )

            column = 2

            for day in dates:
                day_entries = entries_by_member_and_date.get(
                    (membership.id, day),
                    [],
                )

                if day_entries:
                    sheet.cell(row=row, column=column, value=1)
                    sheet.cell(
                        row=row,
                        column=column + 1,
                        value=self.get_blocks_text(day_entries),
                    )
                else:
                    sheet.cell(row=row, column=column, value=0)
                    sheet.cell(row=row, column=column + 1, value="-")

                column += 2

            row += 1

    def build_day_sheet(
        self,
        sheet,
        day,
        memberships,
        entries_by_member_and_date,
        people_count,
    ):
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=6,
        )
        sheet.cell(row=1, column=1, value=f"График на {day.strftime('%d.%m.%Y')}")

        sheet.merge_cells(
            start_row=3,
            start_column=1,
            end_row=3,
            end_column=3,
        )
        sheet.cell(row=3, column=1, value="Все участники")

        sheet.merge_cells(
            start_row=3,
            start_column=5,
            end_row=3,
            end_column=6,
        )
        sheet.cell(row=3, column=5, value="Назначены на смену")

        sheet.cell(row=4, column=1, value="ФИО")
        sheet.cell(row=4, column=2, value="Блок")
        sheet.cell(row=4, column=3, value="Выхожу")

        sheet.cell(row=4, column=5, value="ФИО")
        sheet.cell(row=4, column=6, value="Блок")

        row_all = 5
        row_assigned = 5

        for membership in memberships:
            day_entries = entries_by_member_and_date.get(
                (membership.id, day),
                [],
            )

            member_name = self.get_member_name(membership)

            if day_entries:
                is_assigned = 1
                block_text = self.get_blocks_text(day_entries)
            else:
                is_assigned = 0
                block_text = "-"

            sheet.cell(row=row_all, column=1, value=member_name)
            sheet.cell(row=row_all, column=2, value=block_text)
            sheet.cell(row=row_all, column=3, value=is_assigned)

            if is_assigned:
                sheet.cell(row=row_assigned, column=5, value=member_name)
                sheet.cell(row=row_assigned, column=6, value=block_text)
                row_assigned += 1

            row_all += 1

        self.apply_day_sheet_styles(sheet)

    def get_member_name(self, membership):
        user = membership.user

        parts = [
            user.last_name,
            user.first_name,
            getattr(user, "middle_name", ""),
        ]

        full_name = " ".join(part for part in parts if part).strip()

        return full_name or user.email or str(user)

    def get_blocks_text(self, entries):
        blocks = []

        for entry in entries:
            block_name = entry.work_block.name if entry.work_block else "-"

            if block_name not in blocks:
                blocks.append(block_name)

        return ", ".join(blocks) if blocks else "-"

    def get_safe_sheet_title(self, value):
        forbidden_chars = ['\\', '/', '*', '[', ']', ':', '?']

        title = value

        for char in forbidden_chars:
            title = title.replace(char, "_")

        return title[:31] or "Лист"

    def apply_summary_styles(self, sheet, dates):
        header_fill = PatternFill("solid", fgColor="E8EEF8")
        count_fill = PatternFill("solid", fgColor="F8FAFC")
        thin_side = Side(style="thin", color="D1D5DB")
        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        last_column = 1 + len(dates) * 2
        last_row = sheet.max_row

        for row in range(1, last_row + 1):
            for column in range(1, last_column + 1):
                cell = sheet.cell(row=row, column=column)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

        for row in range(1, 5):
            for column in range(1, last_column + 1):
                cell = sheet.cell(row=row, column=column)
                cell.fill = header_fill
                cell.font = Font(bold=True)

        for column in range(2, last_column + 1):
            sheet.cell(row=3, column=column).fill = count_fill

        for row in range(5, last_row + 1):
            sheet.cell(row=row, column=1).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

        sheet.column_dimensions["A"].width = 34

        for column in range(2, last_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 14

        sheet.row_dimensions[1].height = 22
        sheet.row_dimensions[2].height = 22
        sheet.row_dimensions[3].height = 22
        sheet.row_dimensions[4].height = 26

        for row in range(5, last_row + 1):
            sheet.row_dimensions[row].height = 32

        sheet.freeze_panes = "B5"

    def apply_day_sheet_styles(self, sheet):
        header_fill = PatternFill("solid", fgColor="E8EEF8")
        subheader_fill = PatternFill("solid", fgColor="F8FAFC")
        thin_side = Side(style="thin", color="D1D5DB")
        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        max_row = sheet.max_row

        for row in range(1, max_row + 1):
            for column in range(1, 7):
                cell = sheet.cell(row=row, column=column)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

                if column != 4:
                    cell.border = border

        for column in range(1, 7):
            if column == 4:
                continue

            sheet.cell(row=1, column=column).fill = header_fill
            sheet.cell(row=1, column=column).font = Font(bold=True)

        for column in range(1, 4):
            sheet.cell(row=3, column=column).fill = subheader_fill
            sheet.cell(row=3, column=column).font = Font(bold=True)

            sheet.cell(row=4, column=column).fill = header_fill
            sheet.cell(row=4, column=column).font = Font(bold=True)

        for column in range(5, 7):
            sheet.cell(row=3, column=column).fill = subheader_fill
            sheet.cell(row=3, column=column).font = Font(bold=True)

            sheet.cell(row=4, column=column).fill = header_fill
            sheet.cell(row=4, column=column).font = Font(bold=True)

        for row in range(5, max_row + 1):
            sheet.cell(row=row, column=1).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

            sheet.cell(row=row, column=5).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 28
        sheet.column_dimensions["C"].width = 12
        sheet.column_dimensions["D"].width = 4
        sheet.column_dimensions["E"].width = 34
        sheet.column_dimensions["F"].width = 28

        sheet.row_dimensions[1].height = 26
        sheet.row_dimensions[3].height = 24
        sheet.row_dimensions[4].height = 24

        for row in range(5, max_row + 1):
            sheet.row_dimensions[row].height = 24

        sheet.freeze_panes = "A5"

class ScheduleEditDataView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        schedule = get_object_or_404(
            Schedule.objects.select_related(
                "squad",
                "availability_form",
            ),
            pk=pk,
        )

        if not can_manage_roster(request.user, schedule.squad):
            raise PermissionDenied("Недостаточно прав для редактирования графика.")

        selected_date = self.get_selected_date(request, schedule)

        days = self.get_schedule_days(schedule)

        needs = (
            schedule.needs
            .filter(date=selected_date)
            .select_related("work_block")
            .order_by("starts_at", "ends_at", "work_block__name", "id")
        )

        entries = (
            schedule.entries
            .filter(date=selected_date)
            .select_related(
                "membership",
                "membership__user",
                "work_block",
                "need",
            )
            .order_by(
                "starts_at",
                "work_block__name",
                "membership__user__last_name",
                "membership__user__first_name",
            )
        )

        entries_by_need = {}

        for entry in entries:
            entries_by_need.setdefault(entry.need_id, []).append(entry)

        response_needs = []

        for need in needs:
            assigned_entries = entries_by_need.get(need.id, [])

            response_needs.append({
                "id": need.id,
                "date": need.date,
                "work_block": need.work_block_id,
                "work_block_name": need.work_block.name if need.work_block else "",
                "starts_at": need.starts_at,
                "ends_at": need.ends_at,
                "required_people": need.required_people,
                "assigned_count": len(assigned_entries),
                "shortage": max(need.required_people - len(assigned_entries), 0),
                "entries": [
                    self.serialize_entry(entry)
                    for entry in assigned_entries
                ],
                "candidates": self.get_candidates_for_need(schedule, need),
            })

        return Response({
            "schedule": {
                "id": schedule.id,
                "title": schedule.title,
                "status": schedule.status,
                "squad": schedule.squad_id,
                "squad_name": schedule.squad.name,
                "availability_form": schedule.availability_form_id,
                "period_start": schedule.period_start,
                "period_end": schedule.period_end,
            },
            "days": days,
            "selected_date": selected_date,
            "needs": response_needs,
        })

    def get_selected_date(self, request, schedule):
        raw_date = request.query_params.get("date")

        if raw_date:
            try:
                selected_date = date.fromisoformat(raw_date)
            except ValueError:
                raise ValidationError({
                    "date": "Некорректная дата. Используйте формат YYYY-MM-DD."
                })

            if not schedule.period_start <= selected_date <= schedule.period_end:
                raise ValidationError({
                    "date": "Дата выходит за пределы периода графика."
                })

            return selected_date

        return schedule.period_start

    def get_schedule_days(self, schedule):
        days = []
        current = schedule.period_start

        while current <= schedule.period_end:
            days.append(current)
            current += timedelta(days=1)

        return days

    def serialize_entry(self, entry):
        return {
            "id": entry.id,
            "membership": entry.membership_id,
            "member_name": self.get_member_name(entry.membership),
            "work_block": entry.work_block_id,
            "work_block_name": entry.work_block.name if entry.work_block else "",
            "date": entry.date,
            "starts_at": entry.starts_at,
            "ends_at": entry.ends_at,
            "status": entry.status,
        }

    def get_candidates_for_need(self, schedule, need):
        if not schedule.availability_form_id:
            return []

        assigned_membership_ids = set(
            schedule.entries
            .filter(need=need)
            .values_list("membership_id", flat=True)
        )

        busy_membership_ids = set(
            schedule.entries
            .filter(
                date=need.date,
                starts_at__lt=need.ends_at,
                ends_at__gt=need.starts_at,
            )
            .exclude(need=need)
            .values_list("membership_id", flat=True)
        )

        slots = (
            AvailabilitySlot.objects
            .filter(
                shift__day__form=schedule.availability_form,
                shift__day__date=need.date,
                shift__starts_at__lte=need.starts_at,
                shift__ends_at__gte=need.ends_at,
                is_available=True,
                membership__squad=schedule.squad,
                membership__is_active=True,
            )
            .select_related(
                "membership",
                "membership__user",
            )
            .order_by(
                "membership__user__last_name",
                "membership__user__first_name",
            )
        )

        candidates = []

        seen_memberships = set()

        for slot in slots:
            membership = slot.membership

            if membership.id in seen_memberships:
                continue

            seen_memberships.add(membership.id)

            candidates.append({
                "membership": membership.id,
                "member_name": self.get_member_name(membership),
                "is_assigned": membership.id in assigned_membership_ids,
                "is_busy": membership.id in busy_membership_ids,
            })

        return candidates

    def get_member_name(self, membership):
        if not membership:
            return ""

        user = membership.user

        parts = [
            user.last_name,
            user.first_name,
            getattr(user, "middle_name", ""),
        ]

        full_name = " ".join(part for part in parts if part).strip()

        return full_name or user.email or str(user)